from django.shortcuts import render, redirect, get_object_or_404
from .models import Ingreso, Gasto
from .forms import IngresoForm, GastoForm
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Image as RLImage
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.units import cm
from django.conf import settings
from pathlib import Path
from django.db.models import Sum
from calendar import monthrange

FILTER_SESSION_KEY = 'listado_filtros'


def _rango_mes_actual():
    hoy = timezone.localdate()
    inicio = hoy.replace(day=1)
    fin = hoy.replace(day=monthrange(hoy.year, hoy.month)[1])
    return inicio.isoformat(), fin.isoformat()


def _guardar_filtros(request, fecha_desde, fecha_hasta, concepto):
    request.session[FILTER_SESSION_KEY] = {
        'desde': fecha_desde or '',
        'hasta': fecha_hasta or '',
        'concepto': concepto or '',
    }


def _filtros_periodo(request):
    # Limpiar filtros compartidos
    if request.GET.get('limpiar') == '1':
        _guardar_filtros(request, '', '', '')
        return '', '', ''

    # Si vienen filtros en la URL (formulario o paginación), se guardan y comparten
    if any(key in request.GET for key in ('desde', 'hasta', 'concepto')):
        fecha_desde = request.GET.get('desde') or ''
        fecha_hasta = request.GET.get('hasta') or ''
        concepto = (request.GET.get('concepto') or '').strip()
        _guardar_filtros(request, fecha_desde, fecha_hasta, concepto)
        return fecha_desde, fecha_hasta, concepto

    # Reutilizar filtros de la otra sección
    guardados = request.session.get(FILTER_SESSION_KEY)
    if guardados is not None:
        return (
            guardados.get('desde', ''),
            guardados.get('hasta', ''),
            guardados.get('concepto', ''),
        )

    # Primera visita: mes actual
    fecha_desde, fecha_hasta = _rango_mes_actual()
    _guardar_filtros(request, fecha_desde, fecha_hasta, '')
    return fecha_desde, fecha_hasta, ''


def _aplicar_filtros(queryset, fecha_desde, fecha_hasta, concepto):
    if fecha_desde:
        queryset = queryset.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        queryset = queryset.filter(fecha__lte=fecha_hasta)
    if concepto:
        queryset = queryset.filter(concepto__icontains=concepto)
    return queryset


@login_required
def lista_ingresos(request):
    ingresos = Ingreso.objects.filter(user=request.user).order_by('-fecha')
    fecha_desde, fecha_hasta, concepto = _filtros_periodo(request)
    ingresos = _aplicar_filtros(ingresos, fecha_desde, fecha_hasta, concepto)

    total_filtrado = ingresos.aggregate(total=Sum('cantidad'))['total'] or 0
    page_obj = Paginator(ingresos, 20).get_page(request.GET.get('page'))

    return render(request, 'finanzas/lista_ingresos.html', {
        'ingresos': page_obj,
        'page_obj': page_obj,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'concepto': concepto,
        'total_filtrado': total_filtrado,
    })


@login_required
def lista_gastos(request):
    gastos = Gasto.objects.filter(user=request.user).order_by('-fecha')
    fecha_desde, fecha_hasta, concepto = _filtros_periodo(request)
    gastos = _aplicar_filtros(gastos, fecha_desde, fecha_hasta, concepto)

    total_filtrado = gastos.aggregate(total=Sum('cantidad'))['total'] or 0
    page_obj = Paginator(gastos, 20).get_page(request.GET.get('page'))

    return render(request, 'finanzas/lista_gastos.html', {
        'gastos': page_obj,
        'page_obj': page_obj,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'concepto': concepto,
        'total_filtrado': total_filtrado,
    })

@login_required
def añadir_ingreso(request):
    if request.method == 'POST':
        form = IngresoForm(request.POST)
        if form.is_valid():
            ingreso = form.save(commit=False)
            ingreso.user = request.user
            ingreso.save()
            messages.success(request, 'Ingreso guardado correctamente.')
            return redirect('lista_ingresos')
        messages.error(request, 'No se pudo guardar el ingreso. Revisa los datos.')
        return redirect('lista_ingresos')
    return redirect('lista_ingresos')

@login_required
def añadir_gasto(request):
    if request.method == 'POST':
        form = GastoForm(request.POST)
        if form.is_valid():
            gasto = form.save(commit=False)
            gasto.user = request.user
            gasto.save()
            messages.success(request, 'Gasto guardado correctamente.')
            return redirect('lista_gastos')
        messages.error(request, 'No se pudo guardar el gasto. Revisa los datos.')
        return redirect('lista_gastos')
    return redirect('lista_gastos')

@login_required
def editar_ingreso(request, pk):
    ingreso = get_object_or_404(Ingreso, pk=pk, user=request.user)
    if request.method == 'POST':
        form = IngresoForm(request.POST, instance=ingreso)
        if form.is_valid():
            ingreso = form.save(commit=False)
            ingreso.user = request.user
            ingreso.save()
            messages.success(request, 'Ingreso actualizado correctamente.')
            return redirect('lista_ingresos')
        messages.error(request, 'No se pudo actualizar el ingreso. Revisa los datos.')
        return redirect('lista_ingresos')
    return redirect('lista_ingresos')

@login_required
def editar_gasto(request, pk):
    gasto = get_object_or_404(Gasto, pk=pk, user=request.user)
    if request.method == 'POST':
        form = GastoForm(request.POST, instance=gasto)
        if form.is_valid():
            gasto = form.save(commit=False)
            gasto.user = request.user
            gasto.save()
            messages.success(request, 'Gasto actualizado correctamente.')
            return redirect('lista_gastos')
        messages.error(request, 'No se pudo actualizar el gasto. Revisa los datos.')
        return redirect('lista_gastos')
    return redirect('lista_gastos')

@login_required
def eliminar_ingreso(request, pk):
    ingreso = get_object_or_404(Ingreso, pk=pk, user=request.user)
    ingreso.delete()
    return redirect('lista_ingresos')

@login_required
def eliminar_gasto(request, pk):
    gasto = get_object_or_404(Gasto, pk=pk, user=request.user)
    gasto.delete()
    return redirect('lista_gastos')

@login_required
def generar_seguimiento(request):
    fecha_inicio = parse_date(request.GET.get('fecha_inicio'))
    fecha_fin = parse_date(request.GET.get('fecha_fin'))
    formato = request.GET.get('formato')  # 'excel' o 'pdf'

    if isinstance(fecha_inicio, str):
        fecha_inicio = parse_date(fecha_inicio)
    if isinstance(fecha_fin, str):
        fecha_fin = parse_date(fecha_fin)

    if not fecha_inicio or not fecha_fin:
        return HttpResponse("Fechas inválidas", status=400)

    ingresos = Ingreso.objects.filter(user=request.user, fecha__range=(fecha_inicio, fecha_fin)).order_by('fecha')
    gastos = Gasto.objects.filter(user=request.user, fecha__range=(fecha_inicio, fecha_fin)).order_by('fecha')

    total_ingresos = sum(i.cantidad for i in ingresos)
    total_gastos = sum(g.cantidad for g in gastos)
    saldo_periodo = total_ingresos - total_gastos

    total_ingresos_global = Ingreso.objects.filter(user=request.user).aggregate(Sum('cantidad'))['cantidad__sum'] or 0
    total_gastos_global = Gasto.objects.filter(user=request.user).aggregate(Sum('cantidad'))['cantidad__sum'] or 0
    saldo_actual = total_ingresos_global - total_gastos_global

    fecha_ini_str = fecha_inicio.strftime('%d-%m-%Y')
    fecha_fin_str = fecha_fin.strftime('%d-%m-%Y')
    nombre_archivo = f"ODA - {fecha_ini_str} - {fecha_fin_str}"

    if formato == 'excel':
        wb = Workbook()
        ws_ingresos = wb.active
        ws_ingresos.title = "Ingresos"

        cabecera_fill = PatternFill(start_color="13241E", end_color="13241E", fill_type="solid")
        header_font = Font(bold=True, color="F4F8F5")
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='C8D5CE'),
            right=Side(style='thin', color='C8D5CE'),
            top=Side(style='thin', color='C8D5CE'),
            bottom=Side(style='thin', color='C8D5CE'),
        )

        def estilizar_cabecera(ws):
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = cabecera_fill
                cell.alignment = center_align
                cell.border = thin_border

        def estilizar_cuerpo(ws):
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = thin_border

        ws_ingresos.append(["Fecha", "Concepto", "Cantidad (€)"])
        estilizar_cabecera(ws_ingresos)
        for ingreso in ingresos:
            ws_ingresos.append([
                ingreso.fecha.strftime('%d/%m/%Y'),
                ingreso.concepto,
                float(ingreso.cantidad),
            ])
        estilizar_cuerpo(ws_ingresos)
        ws_ingresos.column_dimensions['A'].width = 14
        ws_ingresos.column_dimensions['B'].width = 45
        ws_ingresos.column_dimensions['C'].width = 14

        ws_gastos = wb.create_sheet("Gastos")
        ws_gastos.append(["Fecha", "Concepto", "Cantidad (€)"])
        estilizar_cabecera(ws_gastos)
        for gasto in gastos:
            ws_gastos.append([
                gasto.fecha.strftime('%d/%m/%Y'),
                gasto.concepto,
                float(gasto.cantidad),
            ])
        estilizar_cuerpo(ws_gastos)
        ws_gastos.column_dimensions['A'].width = 14
        ws_gastos.column_dimensions['B'].width = 45
        ws_gastos.column_dimensions['C'].width = 14

        ws_totales = wb.create_sheet("Resumen", 0)
        ws_totales.append(["Concepto", "Importe (€)"])
        estilizar_cabecera(ws_totales)
        ws_totales.append(["Total ingresos (periodo)", float(total_ingresos)])
        ws_totales.append(["Total gastos (periodo)", float(total_gastos)])
        ws_totales.append(["Saldo del periodo", float(saldo_periodo)])
        ws_totales.append(["Saldo actual", float(saldo_actual)])
        ws_totales.append(["", ""])
        ws_totales.append([
            "Periodo",
            f"{fecha_inicio.strftime('%d/%m/%Y')} — {fecha_fin.strftime('%d/%m/%Y')}",
        ])
        estilizar_cuerpo(ws_totales)
        for row in ws_totales.iter_rows(min_row=2, max_row=5, min_col=1, max_col=1):
            for cell in row:
                cell.font = bold_font
        ws_totales['A7'].font = bold_font
        ws_totales['B7'].alignment = Alignment(horizontal='right', vertical='center')
        ws_totales.column_dimensions['A'].width = 28
        ws_totales.column_dimensions['B'].width = 28

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename={nombre_archivo}.xlsx'
        wb.save(response)
        return response

    elif formato == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.pdf"'

        ink = colors.HexColor('#13241e')
        moss = colors.HexColor('#2a5c48')
        foam = colors.HexColor('#f4f8f5')
        positive = colors.HexColor('#1f6b4a')
        negative = colors.HexColor('#9b3a2f')
        line = colors.HexColor('#c8d5ce')
        row_alt = colors.HexColor('#e8f0eb')

        doc = SimpleDocTemplate(
            response,
            pagesize=A4,
            leftMargin=1.6 * cm,
            rightMargin=1.6 * cm,
            topMargin=1.6 * cm,
            bottomMargin=1.6 * cm,
        )
        styles = getSampleStyleSheet()
        style_brand = ParagraphStyle(
            'OdaBrand',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=24,
            leading=28,
            textColor=ink,
            alignment=TA_LEFT,
            spaceBefore=0,
            spaceAfter=0,
        )
        style_subtitle = ParagraphStyle(
            'OdaSubtitle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=11,
            leading=14,
            textColor=moss,
            alignment=TA_CENTER,
            spaceBefore=10,
            spaceAfter=6,
        )
        style_meta = ParagraphStyle(
            'OdaMeta',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            leading=12,
            textColor=ink,
            alignment=TA_CENTER,
            spaceBefore=0,
            spaceAfter=12,
        )
        style_section = ParagraphStyle(
            'OdaSection',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=12,
            textColor=ink,
            spaceBefore=8,
            spaceAfter=8,
        )
        style_cell = ParagraphStyle(
            'OdaCell',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            textColor=ink,
            leading=12,
        )
        style_empty = ParagraphStyle(
            'OdaEmpty',
            parent=styles['Normal'],
            fontName='Helvetica-Oblique',
            fontSize=9,
            textColor=colors.HexColor('#6b7f75'),
            alignment=TA_CENTER,
        )

        def money_text(value):
            return f"{float(value):,.2f} €".replace(',', 'X').replace('.', ',').replace('X', '.')

        def build_movement_table(rows, header_color):
            data = [[
                Paragraph('Fecha', style_cell),
                Paragraph('Concepto', style_cell),
                Paragraph('Cantidad', style_cell),
            ]]
            rows = list(rows)
            if rows:
                for item in rows:
                    signo = '+' if header_color == positive else '−'
                    data.append([
                        Paragraph(item.fecha.strftime('%d/%m/%Y'), style_cell),
                        Paragraph(item.concepto.replace('\n', '<br/>'), style_cell),
                        Paragraph(f"{signo}{money_text(item.cantidad)}", style_cell),
                    ])
            else:
                data.append([
                    Paragraph('—', style_empty),
                    Paragraph('Sin movimientos en este periodo', style_empty),
                    Paragraph('—', style_empty),
                ])

            table = Table(data, colWidths=[2.4 * cm, 11.2 * cm, 3.2 * cm], hAlign='CENTER')
            style_commands = [
                ('BACKGROUND', (0, 0), (-1, 0), header_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), foam),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 7),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
                ('BOX', (0, 0), (-1, -1), 0.6, line),
                ('LINEBELOW', (0, 0), (-1, -2), 0.4, line),
            ]
            for i in range(1, len(data)):
                if i % 2 == 0:
                    style_commands.append(('BACKGROUND', (0, i), (-1, i), row_alt))
            table.setStyle(TableStyle(style_commands))
            return table, len(rows)

        saldo_periodo_color = '#1f6b4a' if saldo_periodo >= 0 else '#9b3a2f'
        saldo_actual_color = '#1f6b4a' if saldo_actual >= 0 else '#9b3a2f'
        card_style = ParagraphStyle('OdaCard', parent=styles['Normal'], alignment=TA_CENTER, leading=14)

        summary = Table(
            [[
                Paragraph(
                    f'<font size="8" color="#5a6f66">INGRESOS PERIODO</font><br/>'
                    f'<font size="12" color="#1f6b4a"><b>{money_text(total_ingresos)}</b></font>',
                    card_style,
                ),
                Paragraph(
                    f'<font size="8" color="#5a6f66">GASTOS PERIODO</font><br/>'
                    f'<font size="12" color="#9b3a2f"><b>{money_text(total_gastos)}</b></font>',
                    card_style,
                ),
                Paragraph(
                    f'<font size="8" color="#5a6f66">SALDO PERIODO</font><br/>'
                    f'<font size="12" color="{saldo_periodo_color}"><b>{money_text(saldo_periodo)}</b></font>',
                    card_style,
                ),
                Paragraph(
                    f'<font size="8" color="#5a6f66">SALDO ACTUAL</font><br/>'
                    f'<font size="12" color="{saldo_actual_color}"><b>{money_text(saldo_actual)}</b></font>',
                    card_style,
                ),
            ]],
            colWidths=[4.2 * cm, 4.2 * cm, 4.2 * cm, 4.2 * cm],
            hAlign='CENTER',
        )
        summary.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), foam),
            ('BOX', (0, 0), (-1, 0), 0.8, line),
            ('INNERGRID', (0, 0), (-1, 0), 0.5, line),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))

        tabla_ingresos, n_ingresos = build_movement_table(ingresos, positive)
        tabla_gastos, n_gastos = build_movement_table(gastos, negative)

        logo_path = Path(settings.BASE_DIR) / 'finanzas' / 'static' / 'finanzas' / 'img' / 'favicon.png'
        logo = RLImage(str(logo_path), width=1.15 * cm, height=1.15 * cm)
        header_brand = Table(
            [[logo, Paragraph('ODA', style_brand)]],
            colWidths=[1.4 * cm, 3.2 * cm],
            hAlign='CENTER',
        )
        header_brand.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
            ('ALIGN', (1, 0), (1, 0), 'LEFT'),
        ]))

        elements = [
            header_brand,
            Spacer(1, 8),
            Paragraph('Resumen de seguimiento', style_subtitle),
            Paragraph(
                f"Periodo: <b>{fecha_inicio.strftime('%d/%m/%Y')}</b> — <b>{fecha_fin.strftime('%d/%m/%Y')}</b>",
                style_meta,
            ),
            HRFlowable(width='100%', thickness=1, color=moss, spaceBefore=4, spaceAfter=14),
            summary,
            Spacer(1, 18),
            Paragraph(f'Ingresos ({n_ingresos})', style_section),
            tabla_ingresos,
            Spacer(1, 16),
            Paragraph(f'Gastos ({n_gastos})', style_section),
            tabla_gastos,
        ]

        doc.build(elements)
        return response

    return HttpResponse("Formato no válido", status=400)